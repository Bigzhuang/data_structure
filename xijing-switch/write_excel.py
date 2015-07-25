#-*-coding:utf-8-*-
from openpyxl import Workbook
wb=Workbook()
ws = wb.active
n=1
f=open('makefile/port.txt','r')


ws['A1']='交换机'
ws['B1']='端口'
ws['C1']='MAC地址'
ws['D1']='IP地址'
ws['E1']='vlan'
for line in f.readlines():
    line1=line.split('#')
    if len(line1)==2 and line1[1]=='show por add\r\n':
        print line1
        n+=1
        ws['A'+str(n)]=line1[0]
    line=line.split()
    if line!=[] and line[0].isdigit():
        
        ws['B'+str(n)]=line[3]
        ws['C'+str(n)]=line[1]
        ws['E'+str(n)]=line[0]
        with open('makefile/arp.txt','r')as g:
            for hang in g.readlines():
                hang=hang.split()
                if hang!=[] and hang[0]=='Internet':
                    if hang[3]==line[1]:
                        ws['D'+str(n)]=hang[1]
        n+=1

f.close()
wb.save('makefile/output.xlsx')    