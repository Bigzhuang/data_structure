#-*-coding:utf-8-*-
from openpyxl import Workbook
wb=Workbook()
ws = wb.active
n=1
f=open('c:\Users\Dazhuang\Desktop\show_port','r')


ws['A1']='端口'
ws['B1']='mac地址'
ws['C1']='IP地址'
ws['D1']='vlan'
for line in f.readlines():
    line=line.split()
    if line!=[] and line[0].isdigit():
        n+=1
        ws['A'+str(n)]=line[3]
        ws['B'+str(n)]=line[1]
        ws['D'+str(n)]=line[0]
        with open('c:\Users\Dazhuang\Desktop\show_arp','r')as g:
            for hang in g.readlines():
                hang=hang.split()
                if hang!=[] and hang[0]=='Internet':
                    if hang[3]==line[1]:
                        ws['C'+str(n)]=hang[1]

f.close()
wb.save('hehe.xlsx')    